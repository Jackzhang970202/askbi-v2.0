"""
ReportRunner - 报表生成协调器（逐列处理模式）

工作流程:
  Phase 0: 准备 - 查找数据文件, 构建元数据
  Phase 1: 创建带表头的空 Excel
  Phase 2: 解析智能体分析报表规则, 生成每列处理说明
  Phase 3: 逐列处理 -
      对每一列:
        a. 专家智能体生成计算代码（不写 Excel）
        b. 代码执行器执行代码
        c. 若出错 → 根据错误重新生成（自动重试）
        d. 验证智能体验证输出数据
        e. 若不通过 → 根据反馈重新生成
        f. 验证通过 → 将验证后的数据写入 Excel
  Phase 4: 返回 Excel 文件路径
"""

import os
import re
import json
import datetime
import time
import traceback
import logging
import asyncio
import pandas as pd
from typing import List, Optional, Dict, Any, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from utils.db_utils import db_utils

logger = logging.getLogger(__name__)


class ReportRunner:
    """报表生成专用协调器 - 逐列处理模式"""

    def __init__(self, model_client, user_id: Optional[int] = None, progress_hook=None):
        """
        初始化报表协调器

        Args:
            model_client: 大模型客户端
            user_id: 用户ID
            progress_hook: 进度回调函数 (text: str) -> None
        """
        self.model_client = model_client
        self.user_id = user_id
        self.progress_hook = progress_hook

        # 工作目录
        self.session_dir = os.getcwd()

        # Python 环境配置
        cwd = os.getcwd()
        current_pp = os.environ.get("PYTHONPATH", "")
        if cwd not in current_pp:
            os.environ["PYTHONPATH"] = cwd + os.pathsep + current_pp
        os.environ['PYTHONIOENCODING'] = 'utf-8'
        os.environ['PYTHONUNBUFFERED'] = '1'

    def emit(self, text: str):
        """推送进度到前端"""
        if self.progress_hook:
            try:
                self.progress_hook(str(text).strip())
            except Exception:
                pass

    # ==================== 工具方法 ====================

    def _extract_agent_reply(self, result) -> str:
        """
        从 agent.run() 的 TaskResult 中提取智能体的回复文本。

        Args:
            result: agent.run() 返回的 TaskResult

        Returns:
            智能体回复的文本内容
        """
        if hasattr(result, 'messages'):
            for m in result.messages:
                if hasattr(m, 'content') and isinstance(m.content, str):
                    if getattr(m, 'source', '') != 'user':
                        return m.content
        return str(result)

    @staticmethod
    def _extract_python_code(text: str) -> Optional[str]:
        """
        从文本中提取最后一个 ```python 代码块的内容。
        取最后一个是因为: 如果智能体先写了解释性代码片段, 最后一个才是完整脚本。

        Args:
            text: 包含代码块的文本

        Returns:
            提取的 Python 代码, 如果没找到返回 None
        """
        pattern = r'```python\s*\n(.*?)```'
        matches = re.findall(pattern, text, re.DOTALL)
        if matches:
            return matches[-1].strip()
        # 兜底: 尝试无语言标识的代码块
        pattern2 = r'```\s*\n(.*?)```'
        matches2 = re.findall(pattern2, text, re.DOTALL)
        if matches2:
            for m in reversed(matches2):
                if 'import ' in m and ('pandas' in m or 'json' in m):
                    return m.strip()
        return None

    async def _execute_code_directly(self, code: str) -> Tuple[str, bool]:
        """
        直接执行 Python 代码（绕过 CodeExecutorAgent, 彻底避免代码块污染）。
        将代码写入临时 .py 文件, 用 asyncio 子进程执行。

        Args:
            code: 要执行的完整 Python 代码

        Returns:
            (output, success) 元组
        """
        temp_py = os.path.join(self.session_dir, '_report_exec_temp.py')

        with open(temp_py, 'w', encoding='utf-8') as f:
            f.write(code)

        try:
            proc = await asyncio.create_subprocess_exec(
                'python', temp_py,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                cwd=self.session_dir
            )
            try:
                stdout, stderr = await asyncio.wait_for(
                    proc.communicate(), timeout=120
                )
            except asyncio.TimeoutError:
                proc.kill()
                await proc.communicate()
                return "执行超时(120秒)", False

            out_text = stdout.decode('utf-8', errors='replace')
            err_text = stderr.decode('utf-8', errors='replace')
            combined = out_text
            if err_text.strip():
                combined += "\n[STDERR]\n" + err_text

            return combined, (proc.returncode == 0)

        except Exception as e:
            return f"执行异常: {str(e)}\n{traceback.format_exc()}", False
        finally:
            try:
                if os.path.exists(temp_py):
                    os.remove(temp_py)
            except OSError:
                pass

    # ==================== Phase 0: 准备阶段 ====================

    def _find_data_files(self, chat_id: str) -> List[str]:
        """
        查找会话的数据文件（优先使用拆分后的 sheet 文件）

        Args:
            chat_id: 会话ID

        Returns:
            Excel 文件路径列表
        """
        session = db_utils.get_chat_session(chat_id)
        if not session:
            raise Exception(f"会话不存在: {chat_id}")

        user_id = self.user_id or session.get('user_id')
        user_prefix = f"user_{user_id}" if user_id else "admin"
        data_dir = os.path.join(self.session_dir, "split_files", user_prefix, chat_id)

        all_files = []
        if os.path.exists(data_dir):
            for f in os.listdir(data_dir):
                if f.endswith(('.xlsx', '.xls', '.csv')) and not f.startswith('~$'):
                    all_files.append(os.path.join(data_dir, f))

        if not all_files:
            raise Exception(f"未找到数据文件: {data_dir}")

        # 优先使用拆分后的 sheet 文件（文件名含 __ 的）
        split_files = [f for f in all_files if '__' in os.path.basename(f)]
        result = split_files if split_files else all_files

        logger.info(f"[报表] 找到 {len(result)} 个数据文件")
        return result

    def _build_file_metadata(self, excel_files: List[str]) -> str:
        """
        构建文件元数据（表头 + 样例数据 + 行数）

        Args:
            excel_files: Excel 文件路径列表

        Returns:
            格式化的元数据字符串
        """
        parts = []
        for i, fp in enumerate(excel_files):
            fname = os.path.basename(fp)
            parts.append(f"\n### 文件 {i + 1}: {fname}")
            parts.append(f"完整路径: {fp}")
            try:
                df = pd.read_excel(fp)
                parts.append(f"表头: {list(df.columns)}")
                parts.append(f"数据行数（不含表头）: {len(df)}")
                sample = df.head(5)
                parts.append(f"样例数据（前5行）:\n{sample.to_string(index=False)}")
            except Exception as e:
                parts.append(f"读取失败: {e}")
        return "\n".join(parts)

    def _create_output_path(self, chat_id: str, rule_name: str) -> str:
        """创建报表输出路径"""
        user_prefix = f"user_{self.user_id}" if self.user_id else "admin"
        report_dir = os.path.join(self.session_dir, "report_files", user_prefix, chat_id)
        os.makedirs(report_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(report_dir, f"{rule_name}_{ts}.xlsx")

    # ==================== Phase 1: 创建表头Excel ====================

    def _create_excel_with_headers(self, headers: List[str], output_path: str):
        """创建只有表头的空 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "报表"

        hfont = Font(bold=True)
        halign = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hfont
            cell.alignment = halign
            cell.border = border

        wb.save(output_path)
        logger.info(f"[报表] 已创建带表头的 Excel: {output_path}")

    # ==================== Phase 2: 解析规则 ====================

    async def _analyze_rules(self, rule: str, headers: List[str],
                             file_metadata: str, file_list: List[str]) -> str:
        """
        调用解析智能体分析报表规则, 生成每列处理说明

        Args:
            rule: 报表规则文本
            headers: 目标表头列表
            file_metadata: 文件元数据
            file_list: 文件路径列表

        Returns:
            完整的分析结果文本
        """
        from agents_new.report_analyzer import create_agent_template as analyzer_tpl
        analyzer = analyzer_tpl(self.model_client)

        file_paths = json.dumps(
            [p.replace(os.sep, '/') for p in file_list],
            ensure_ascii=False
        )

        prompt = f"""请分析以下报表生成规则, 为目标表头的每一列生成详细的处理说明。

## 报表规则
{rule}

## 目标表头（共{len(headers)}列）
{json.dumps(headers, ensure_ascii=False)}

## 原始数据文件信息
{file_metadata}

## 文件路径列表
{file_paths}

请严格按照格式要求, 逐列输出处理说明。每列必须包含: 数据来源、处理方式、计算公式（如需要）、筛选条件（如有）。"""

        self.emit("[report] 解析智能体正在分析报表规则...")
        logger.info("[报表] 开始分析报表规则")

        res = await analyzer.run(task=prompt)

        # 提取分析结果文本
        analysis_text = ""
        if hasattr(res, 'messages'):
            for m in res.messages:
                if hasattr(m, 'content') and isinstance(m.content, str):
                    if getattr(m, 'source', '') != 'user':
                        analysis_text = m.content
        if not analysis_text:
            analysis_text = str(res)

        self.emit("[report] 规则分析完成")
        logger.info(f"[报表] 规则分析完成, 结果长度={len(analysis_text)}")
        return analysis_text

    # ==================== Phase 3: 逐列处理 ====================

    async def _process_single_column(
        self,
        col_idx: int,
        header: str,
        analysis_text: str,
        file_metadata: str,
        file_list: List[str],
        rule: str,
        output_path: str
    ) -> bool:
        """
        处理单列的完整流程（手动编排模式, 不使用 RoundRobinGroupChat）:
          1. 专家生成代码 → 我们提取代码 → 直接执行 → 验证器验证
          2. 如有错误, 传递错误信息给专家重新生成
          3. 验证通过后将数据写入 Excel

        关键改进:
          - 不使用 CodeExecutorAgent (它会从对话历史提取所有代码块拼接执行)
          - 改用 _extract_python_code 只提取最后一个代码块
          - 改用 _execute_code_directly 直接写文件执行, 彻底隔离

        Args:
            col_idx: 列索引（从0开始）
            header: 列名
            analysis_text: 解析智能体的完整分析结果
            file_metadata: 文件元数据
            file_list: 文件路径列表
            rule: 报表规则文本
            output_path: 输出 Excel 路径

        Returns:
            该列是否处理成功
        """
        # 临时 JSON 文件路径（用于存储该列数据）
        temp_file = os.path.join(
            self.session_dir, f"_report_temp_{col_idx}.json"
        ).replace(os.sep, '/')

        # 清理旧临时文件
        if os.path.exists(temp_file):
            os.remove(temp_file)

        self.emit(f"[report] >> 开始处理第{col_idx + 1}列: {header}")
        logger.info(f"[报表] 开始处理第{col_idx + 1}列: {header}")

        # 构建文件路径字符串
        file_paths_str = json.dumps(
            [p.replace(os.sep, '/') for p in file_list],
            ensure_ascii=False
        )

        # 构建初始任务提示词
        initial_task = f"""请为报表的第{col_idx + 1}列「{header}」生成数据处理代码。

## 报表规则
{rule}

## 解析智能体的分析结果（每列处理说明）
{analysis_text}

## 原始数据文件信息
{file_metadata}

## 文件路径列表
{file_paths_str}

## 当前任务
只处理「{header}」这一列的数据, 将结果保存为 JSON 列表。

## 代码结构要求
你生成的 Python 脚本必须严格遵循以下结构:
1. 第一行: import pandas as pd
2. 第二行: import json
3. 定义 FILE_LIST 变量, 值为: {file_paths_str}
4. 用 pd.read_excel() 读取需要的文件
5. 编写「{header}」列的处理逻辑, 将最终结果存入 result 列表
6. 用以下代码保存结果到临时文件:
   with open('{temp_file}', 'w', encoding='utf-8') as f:
       json.dump(result, f, ensure_ascii=False, default=str)
7. 最后两行必须是:
   print("COLUMN_RESULT_COUNT:" + str(len(result)))
   print("COLUMN_PREVIEW:" + str(result[:5]))

## 约束（必须遵守）
- 你的回复中只能包含**一个** ```python 代码块, 这个代码块就是完整脚本
- 第一行必须是 import 语句, 无任何缩进
- 只处理「{header}」列, 不涉及其他列
- **禁止**生成 Excel 文件、**禁止**使用 openpyxl
- **必须**使用英文半角标点符号
- **禁止**使用中文全角标点
- 分析结果中的伪代码仅供参考, 不要直接复制
- 读取文件时使用 FILE_LIST 中的完整路径"""

        # 创建全新的智能体实例（每次新建避免历史污染）
        from agents_new.report_specialist import create_agent_template as spec_tpl
        from agents_new.report_validator import create_agent_template as val_tpl

        MAX_RETRIES = 4
        current_task = initial_task

        for attempt in range(MAX_RETRIES):
            attempt_num = attempt + 1
            self.emit(f"[report]   第{col_idx + 1}列 尝试 {attempt_num}/{MAX_RETRIES}")
            logger.info(f"[报表] 列「{header}」第{attempt_num}次尝试")

            # ---- Step 1: 调用专家智能体生成代码 ----
            specialist = spec_tpl(self.model_client)
            try:
                spec_result = await specialist.run(task=current_task)
                spec_text = self._extract_agent_reply(spec_result)
            except Exception as e:
                logger.error(f"[报表] 专家智能体异常: {e}")
                current_task = f"上次调用出现异常: {e}\n请重新生成完整的 Python 脚本。"
                continue

            logger.info(f"[报表] 专家回复长度={len(spec_text)}")

            # ---- Step 2: 从专家回复中提取 Python 代码 ----
            code = self._extract_python_code(spec_text)
            if not code:
                logger.warning("[报表] 未能从专家回复中提取到 Python 代码")
                current_task = (
                    "你的回复中没有找到 Python 代码块。\n"
                    "请重新生成, 回复中必须包含且仅包含一个 ```python 代码块, "
                    "代码块就是完整的可执行脚本。\n\n"
                    f"原始任务:\n{initial_task}"
                )
                continue

            logger.info(f"[报表] 提取代码长度={len(code)}, 前80字符: {code[:80]}")

            # ---- Step 3: 直接执行代码（不经过 CodeExecutorAgent）----
            exec_output, exec_success = await self._execute_code_directly(code)
            logger.info(
                f"[报表] 代码执行: success={exec_success}, "
                f"输出前200字符: {exec_output[:200]}"
            )

            # ---- Step 4: 调用验证智能体 ----
            validator = val_tpl(self.model_client)
            val_prompt = f"以下是代码执行器的输出, 请验证:\n\n{exec_output}"
            try:
                val_result = await validator.run(task=val_prompt)
                val_text = self._extract_agent_reply(val_result)
            except Exception as e:
                logger.error(f"[报表] 验证智能体异常: {e}")
                val_text = "验证失败: 验证器异常"

            logger.info(f"[报表] 验证结果: {val_text[:200]}")

            # ---- Step 5: 判断验证结果 ----
            if "验证通过" in val_text:
                self.emit(f"[report]   第{col_idx + 1}列 验证通过!")
                logger.info(f"[报表] 列「{header}」第{attempt_num}次验证通过")
                break
            else:
                self.emit(f"[report]   第{col_idx + 1}列 验证未通过, 准备重试...")
                # 构建重试任务: 包含原始任务 + 上次执行结果 + 验证器反馈
                current_task = (
                    f"上次生成的代码执行结果:\n{exec_output}\n\n"
                    f"验证器反馈:\n{val_text}\n\n"
                    f"请根据以上反馈修正代码, 重新生成**完整的** Python 脚本。\n"
                    f"注意: 回复中只能包含一个 ```python 代码块, "
                    f"第一行必须是 import 语句, 无任何缩进。\n\n"
                    f"原始任务提醒:\n"
                    f"- 处理列: {header}\n"
                    f"- 文件列表: {file_paths_str}\n"
                    f"- 临时文件: {temp_file}\n"
                    f"- 必须输出 COLUMN_RESULT_COUNT 和 COLUMN_PREVIEW"
                )
        else:
            # for-else: 所有重试都失败了
            logger.warning(f"[报表] 列「{header}」{MAX_RETRIES}次尝试均失败")
            self.emit(f"[report]   第{col_idx + 1}列 处理失败（已重试{MAX_RETRIES}次）")
            return False

        # ---- 验证通过, 读取临时文件 ----
        if not os.path.exists(temp_file):
            logger.warning(f"[报表] 列「{header}」: 临时文件未生成")
            self.emit(f"[report] 列「{header}」: 数据未生成")
            return False

        try:
            with open(temp_file, 'r', encoding='utf-8') as f:
                column_data = json.load(f)
        except Exception as e:
            logger.error(f"[报表] 读取临时文件失败: {e}")
            self.emit(f"[report] 列「{header}」: 读取数据失败")
            return False
        finally:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except OSError:
                pass

        if not isinstance(column_data, list):
            logger.error(f"[报表] 列「{header}」: 数据格式不是列表")
            return False

        # ---- 写入 Excel ----
        try:
            self._write_column_to_excel(output_path, col_idx, column_data)
            self.emit(
                f"[report] OK 第{col_idx + 1}列「{header}」写入完成 ({len(column_data)}行)"
            )
            logger.info(f"[报表] 列「{header}」写入成功, {len(column_data)}行")
            return True
        except Exception as e:
            logger.error(f"[报表] 写入 Excel 失败: {e}")
            self.emit(f"[report] 列「{header}」写入 Excel 失败: {e}")
            return False

    def _write_column_to_excel(self, path: str, col_idx: int, data: List[Any]):
        """
        将单列验证通过的数据写入 Excel（从第2行开始, 第1行为表头）

        Args:
            path: Excel 文件路径
            col_idx: 列索引（从0开始）
            data: 数据列表
        """
        wb = load_workbook(path)
        ws = wb.active
        for ri, val in enumerate(data, start=2):
            ws.cell(row=ri, column=col_idx + 1).value = val
        wb.save(path)

    # ==================== 主入口 ====================

    async def generate_report(
        self,
        chat_id: str,
        rule_name: str,
        rule: str,
        headers: List[str]
    ) -> str:
        """
        生成报表的主入口

        流程: 准备 -> 创建Excel表头 -> 分析规则 -> 逐列处理 -> 返回路径

        Args:
            chat_id: 会话ID
            rule_name: 报表规则名称
            rule: 报表处理规则文本
            headers: 目标表头列表

        Returns:
            生成的 Excel 文件路径

        Raises:
            Exception: 生成失败时抛出异常
        """
        logger.info(
            f"[报表] 开始生成: {rule_name}, chat_id={chat_id}, 共{len(headers)}列"
        )
        t0 = time.time()

        try:
            # Phase 0: 准备
            self.emit("[report] 准备阶段: 查找数据文件...")
            files = self._find_data_files(chat_id)
            metadata = self._build_file_metadata(files)
            output_path = self._create_output_path(chat_id, rule_name)

            self.emit(f"[report] 找到 {len(files)} 个数据文件")

            # Phase 1: 创建带表头的空 Excel
            self.emit("[report] 创建报表文件（仅表头）...")
            self._create_excel_with_headers(headers, output_path)

            # Phase 2: 解析规则
            self.emit("[report] 解析智能体正在分析规则...")
            analysis = await self._analyze_rules(rule, headers, metadata, files)

            # Phase 3: 逐列处理
            self.emit(f"[report] 开始逐列处理, 共 {len(headers)} 列")
            ok_count = 0
            fail_count = 0

            for i, h in enumerate(headers):
                col_start = time.time()
                success = await self._process_single_column(
                    col_idx=i,
                    header=h,
                    analysis_text=analysis,
                    file_metadata=metadata,
                    file_list=files,
                    rule=rule,
                    output_path=output_path
                )
                col_elapsed = time.time() - col_start

                if success:
                    ok_count += 1
                    logger.info(f"[报表] 第{i+1}列完成, 耗时{col_elapsed:.1f}秒")
                else:
                    fail_count += 1
                    logger.warning(f"[报表] 第{i+1}列失败, 耗时{col_elapsed:.1f}秒")

                self.emit(
                    f"[report] 进度: {i + 1}/{len(headers)} "
                    f"(成功{ok_count}, 失败{fail_count})"
                )

            # 完成
            elapsed = time.time() - t0
            self.emit(
                f"[report] 报表生成完成! "
                f"成功{ok_count}列, 失败{fail_count}列, 耗时{elapsed:.1f}秒"
            )
            logger.info(
                f"[报表] 生成完成: {ok_count}/{len(headers)}列成功, "
                f"耗时{elapsed:.1f}秒, 输出: {output_path}"
            )
            return output_path

        except Exception as e:
            logger.error(f"[报表] 生成失败: {e}\n{traceback.format_exc()}")
            raise Exception(f"报表生成失败: {str(e)}")
