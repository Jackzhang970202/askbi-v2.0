import pandas as pd
import numpy as np
import re
import calendar
from datetime import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
warnings.filterwarnings('ignore')

# ========== 配置：修改此处文件名即可 ==========
DATA_FILE = '考勤模拟数据-二月总和.xlsx'
CARD_MAP_FILE = '项目地出差对应.xlsx'
OUTPUT_FILE = '报表结果.xlsx'

# ========== 读取数据 ==========
df_detail = pd.read_excel(DATA_FILE, sheet_name='考勤明细表')
df_summary = pd.read_excel(DATA_FILE, sheet_name='考勤汇总表')
df_cc_map = pd.read_excel(CARD_MAP_FILE)

# 构建打卡地→考勤地映射（仅用非项目地出差的记录）
# 同时记录项目地出差的打卡地集合
_card_city_map = {}
_card_places_project = set()  # 项目地出差的打卡地
for _, _r in df_cc_map.iterrows():
    if _r['是否项目地出差'] == '否' and pd.notna(_r['考勤地']):
        _card_city_map[_r['打卡地']] = _r['考勤地']
    elif _r['是否项目地出差'] == '是':
        _card_places_project.add(_r['打卡地'])

_sample_date = df_detail['考勤日期'].dropna().iloc[0]
_days_in_month = calendar.monthrange(_sample_date.year, _sample_date.month)[1]

# ========== 筛选在职正式员工和劳务派遣 ==========
# 从明细表获取员工子组信息（汇总表的员工子组列为空）
emp_group_map = df_detail.groupby('人员编号')['员工子组'].first().to_dict()
valid_ids = [eid for eid in df_summary['浪潮工号'].unique()
             if eid in emp_group_map and emp_group_map[eid] in ['正式员工', '劳务派遣人员']]
df_s = df_summary[df_summary['浪潮工号'].isin(valid_ids)].copy()
df_d = df_detail[df_detail['人员编号'].isin(valid_ids)].copy()

_max_ycq = df_s['应出勤天数'].max()
rest_total_days = _days_in_month - _max_ycq

# ========== 考勤数据完整性检查 ==========
_year, _month = _sample_date.year, _sample_date.month
print(f"=== 考勤数据完整性检查（{_year}年{_month}月，共{_days_in_month}天）===")
all_dates = set(pd.date_range(f'{_year}-{_month:02d}-01', periods=_days_in_month, freq='D'))
missing_count = 0
for emp_id in valid_ids:
    emp_dates = set(df_d[df_d['人员编号'] == emp_id]['考勤日期'].dropna())
    missing = sorted(all_dates - emp_dates)
    if missing:
        name = df_s[df_s['浪潮工号'] == emp_id]['职工姓名'].values[0]
        dates_str = ', '.join(d.strftime('%m-%d') for d in missing)
        print(f"  {emp_id} {name}: 缺{len(missing)}天 [{dates_str}]")
        missing_count += 1
if missing_count == 0:
    print(f"  {len(valid_ids)}人数据完整，无缺失")
else:
    print(f"  共{missing_count}人有缺失")
print()

# ========== 辅助函数 ==========
zero_time = time(0, 0)

def is_nonzero_time(t):
    return isinstance(t, time) and t != zero_time

def time_to_str(t):
    if isinstance(t, time):
        return t.strftime('%H:%M:%S')
    return None

def time_str_to_seconds(t_str):
    if pd.isna(t_str) or t_str == '' or t_str is None:
        return None
    t_str = str(t_str).strip()
    try:
        parts = t_str.split(':')
        if len(parts) == 3:
            h, m, s = int(parts[0]), int(parts[1]), int(float(parts[2]))
            return h * 3600 + m * 60 + s
    except:
        return None
    return None

def seconds_to_time_str(seconds):
    if seconds is None or pd.isna(seconds):
        return ''
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"

kq_cities = ['济南', '北京', '呼和浩特', '哈尔滨', '天津', '德州', '成都', '潍坊', '西安', '长沙', '青岛', '齐齐哈尔']

def extract_city_from_card(s):
    if pd.isna(s):
        return None
    s = str(s).strip()
    if not s:
        return None

    # 1. 特殊后缀处理（-呼市、-呼 表示呼和浩特）
    if s.endswith('-呼市') or s.endswith('-呼'):
        return '呼和浩特'

    # 2. 特殊前缀映射
    special_map = {
        '浪潮科技园': '济南', '算谷': '济南', '浪潮内蒙古': '呼和浩特',
        'K济南': '济南', 'K天津': '天津', 'K山东省济南': '济南',
    }
    for prefix, city in special_map.items():
        if s.startswith(prefix):
            return city

    # 3. 优先匹配省/自治区/自治州 + 市格式（最准确）
    # 匹配"xx省xx市"格式
    m = re.search(r'[^省]+省([^市区县]+?市)', s)
    if m:
        return m.group(1).rstrip('市')
    # 匹配"xx自治区xx市"格式
    m = re.search(r'自治区(.+?市)', s)
    if m:
        return m.group(1).rstrip('市')
    # 匹配"xx自治州xx市"格式
    m = re.search(r'自治州(.+?市)', s)
    if m:
        return m.group(1).rstrip('市')

    # 4. 直辖市开头
    for dc in ['北京', '上海', '天津', '重庆']:
        if s.startswith(dc):
            return dc

    # 5. 匹配"xx市"格式（确保市字前面不是数字）
    m = re.search(r'([^0-9]+?市)', s)
    if m:
        city = m.group(1).rstrip('市')
        if len(city) >= 2:
            return city

    # 6. kq_cities匹配（确保是行政区划位置，不是街道名）
    # 只匹配开头或省名后面
    for c in kq_cities:
        # 开头匹配
        if s.startswith(c):
            return c
        # 省名后面匹配
        m = re.search(r'省' + c, s)
        if m:
            return c
        # 自治区后面匹配
        m = re.search(r'自治区' + c, s)
        if m:
            return c

    # 7. 省名后面取2个字作为城市
    provinces = ['河北','山西','辽宁','吉林','黑龙江','江苏','浙江','安徽','福建','江西',
                 '山东','河南','湖北','湖南','广东','海南','四川','贵州','云南','陕西',
                 '甘肃','青海','台湾']
    for prov in provinces:
        if s.startswith(prov):
            rest = s[len(prov):]
            if len(rest) >= 2:
                return rest[:2]

    return None

def card_matches_kq(card_place, kq_city):
    """
    判断打卡地是否匹配考勤地
    返回: True=匹配, False=不匹配(出差), None=不确定(需要人工核实)

    不确定的情况（标黄）：
    1. 打卡地不在映射表中，且城市提取失败
    2. 打卡地不在映射表中，城市提取结果明显不合理（如"省电"、"市区"等）

    注意：以下情况返回False（出差），但也会被标黄让用户核实：
    - 打卡地在映射表中，但映射的城市与考勤地不同
    - 打卡地不在映射表中，城市提取成功但与考勤地不同
    """
    if pd.isna(card_place):
        return True  # 空值视为匹配
    card_place = str(card_place).strip()
    if not card_place:
        return True  # 空字符串视为匹配

    # 查映射表
    if card_place in _card_city_map:
        mapped_city = _card_city_map[card_place]
        if mapped_city == kq_city:
            return True  # 映射表中匹配
        else:
            # 映射表中但城市不同，返回False（出差），但需要人工核实
            return False

    # 不在映射表中，用城市提取
    city = extract_city_from_card(card_place)
    if city is None:
        return None  # 无法识别城市，标黄

    # 检查城市名是否合理（排除明显错误的提取结果）
    invalid_cities = ['省电', '市区', '城区', '县城', '新区', '园区', '科技', '产业', '商业', '住宅', '公寓']
    if city in invalid_cities:
        return None  # 城市名不合理，标黄

    # 不在映射表，城市提取成功，判断是否匹配
    return city == kq_city

qj_cols = ['年休假', '病假', '产检假', '婚假', '丧假', '产假', '产前假', '陪产假', '哺乳假', '事假', '停工留薪假', '子女升学特别假', '育儿假', '子女护理假']

# ========== 补签/补卡还原 ==========
df_restored = df_d.copy()
for idx, row in df_restored.iterrows():
    if is_nonzero_time(row['补签开始时间']):
        bq_str = time_to_str(row['补签开始时间'])
        if bq_str <= '12:00:00':
            df_restored.at[idx, '上班时间'] = bq_str
        else:
            df_restored.at[idx, '下班时间'] = bq_str
    if is_nonzero_time(row['补卡开始时间']):
        bk_str = time_to_str(row['补卡开始时间'])
        if bk_str <= '12:00:00':
            df_restored.at[idx, '上班时间'] = bk_str
        else:
            df_restored.at[idx, '下班时间'] = bk_str

df_workday_orig = df_d[df_d['班次'].isin(['N001', 'N016'])].copy()
df_restday_orig = df_d[df_d['班次'] == '公休日'].copy()
df_workday_r = df_restored[df_restored['班次'].isin(['N001', 'N016'])].copy()

# ========== 逐人计算 ==========
result_list = []

for _, row in df_s.iterrows():
    emp_id = row['浪潮工号']
    ew_orig = df_workday_orig[df_workday_orig['人员编号'] == emp_id]
    er_orig = df_restday_orig[df_restday_orig['人员编号'] == emp_id]
    ew_r = df_workday_r[df_workday_r['人员编号'] == emp_id]

    emp = {}
    emp['序号'] = ''
    emp['浪潮工号'] = emp_id
    emp['职工姓名'] = row['职工姓名']
    emp['所属单位'] = row['所属单位']
    emp['一级部门'] = row['所属部门']
    emp['二级部门'] = row['一级部门']
    emp['岗位'] = row['岗位']
    emp['考勤地'] = row['考勤地']
    emp['应出勤天数'] = row['应出勤天数']

    dk_days = ew_orig[(ew_orig['上班时间'].notna()) & (ew_orig['下班时间'].notna())].shape[0]
    emp['打卡天数'] = dk_days

    cc_days_work = ew_orig[ew_orig['出差合计'] > 0].shape[0]
    emp['出差天数（工作日）'] = cc_days_work

    ycq = row['应出勤天数'] if pd.notna(row['应出勤天数']) else 0
    kq_city = str(row['考勤地']).split('_')[0]
    non_cc = ew_orig[ew_orig['出差合计'] == 0]
    city_mismatch = 0
    need_verify_flag = False  # 是否有需要人工核实的打卡地
    for _, r in non_cc.iterrows():
        sb_ok = card_matches_kq(r['上班卡地'], kq_city)
        # 只检查上班卡地来判断是否出差
        if sb_ok == False:
            city_mismatch += 1
        # 检查是否在映射表中但城市不同（需要人工核实）
        if pd.notna(r['上班卡地']) and str(r['上班卡地']).strip() in _card_city_map:
            if _card_city_map[str(r['上班卡地']).strip()] != kq_city:
                need_verify_flag = True
        if pd.notna(r['下班卡地']) and str(r['下班卡地']).strip() in _card_city_map:
            if _card_city_map[str(r['下班卡地']).strip()] != kq_city:
                need_verify_flag = True
    if ycq > 0:
        cc_rate = round((cc_days_work + city_mismatch) / ycq * 100, 2)  # 百分比，保留两位小数
        emp['出差率（工作日）'] = f"{cc_rate}%" if cc_rate > 0 else ''
    else:
        emp['出差率（工作日）'] = ''
    emp['_出差率不确定'] = need_verify_flag

    cc_days_rest = er_orig[er_orig['出差合计'] > 0].shape[0]
    emp['出差天数（休息日）'] = cc_days_rest

    cd_count = ew_orig[ew_orig['迟到（分钟数）'] > 0].shape[0]
    emp['迟到次数'] = cd_count

    qj_total = ew_orig[qj_cols].sum().sum() if len(ew_orig) > 0 else 0
    qj_days = round(qj_total / 8, 2) if qj_total > 0 else 0
    emp['请假天数\n（工作日）'] = qj_days if qj_days > 0 else ''

    bq_count = row['非工作原因补签卡次数'] if pd.notna(row['非工作原因补签卡次数']) else 0
    emp['补签次数'] = int(bq_count)

    # === 平均上班时间和平均下班时间（新规则）===
    # 规则说明：
    # 1. 下班时间在00:00-04:00之间 -> 设为23:59:59
    # 2. 上班时间在00:00-04:00之间 -> 剔除该上班时间
    # 3. N001班次，下班时间存在且<17:30 -> 下班=17:30
    # 4. N016班次，下班时间存在且<18:00 -> 下班=18:00
    # 5. N001班次，上班时间>17:30且下班为空 -> 上班=08:30, 下班=原上班时间
    # 6. N016班次，上班时间>18:00且下班为空 -> 上班=09:00, 下班=原上班时间
    # 7. N016班次，平均时间减30分钟

    valid_sb = []
    valid_xb = []

    for _, r in ew_r.iterrows():
        sb = r['上班时间']
        xb = r['下班时间']
        bc = r['班次']

        sb_sec = time_str_to_seconds(str(sb)) if pd.notna(sb) and sb != '' else None
        xb_sec = time_str_to_seconds(str(xb)) if pd.notna(xb) and xb != '' else None

        is_n016 = (bc == 'N016')
        is_n001 = (bc == 'N001')

        # 规则1: 下班时间在00:00-04:00之间 -> 设为23:59:59
        if xb_sec is not None and 0 <= xb_sec <= 14400:  # 04:00:00 = 14400秒
            xb_sec = 86399  # 23:59:59

        # 规则2: 上班时间在00:00-04:00之间 -> 剔除
        if sb_sec is not None and 0 <= sb_sec <= 14400:
            sb_sec = None  # 剔除

        # 规则3和4: 下班时间存在且早于阈值
        if xb_sec is not None:
            if is_n001 and xb_sec < 63000:  # 17:30:00 = 63000秒
                xb_sec = 63000
            elif is_n016 and xb_sec < 64800:  # 18:00:00 = 64800秒
                xb_sec = 64800

        # 规则5和6: 上班时间晚于阈值且下班为空 -> 上班重置，下班=原上班时间
        if sb_sec is not None and xb_sec is None:
            if is_n001 and sb_sec > 63000:
                xb_sec = sb_sec  # 下班=原上班时间
                sb_sec = 30600   # 上班=08:30
            elif is_n016 and sb_sec > 64800:
                xb_sec = sb_sec  # 下班=原上班时间
                sb_sec = 32400   # 上班=09:00

        # 收集有效的时间数据
        if sb_sec is not None:
            valid_sb.append(sb_sec)
        if xb_sec is not None:
            valid_xb.append(xb_sec)

    # 计算平均上班时间
    if valid_sb:
        avg_sb = np.mean(valid_sb)
        # 规则8: N016班次减30分钟
        has_n016 = ew_r[ew_r['班次'] == 'N016'].shape[0] > 0
        if has_n016:
            avg_sb -= 1800  # 30分钟 = 1800秒
        emp['平均上班时间'] = seconds_to_time_str(avg_sb)
    else:
        emp['平均上班时间'] = '08:30:00'

    # 计算平均下班时间
    if valid_xb:
        avg_xb = np.mean(valid_xb)
        # 规则8: N016班次减30分钟
        if has_n016:
            avg_xb -= 1800
        emp['平均下班时间'] = seconds_to_time_str(avg_xb)
    else:
        emp['平均下班时间'] = '17:30:00'

    rest_dk = er_orig[er_orig['上班时间'].notna()].shape[0]
    zm_rate = round(rest_dk / rest_total_days * 100, 2) if rest_total_days > 0 else 0  # 百分比，保留两位小数
    emp['周末出勤率'] = f"{zm_rate}%" if zm_rate > 0 else ''

    workday_count = len(ew_orig)
    bh_gt1 = ew_orig[ew_orig['班后贡献'] >= 1].shape[0]
    if workday_count > 0:
        bh_rate = round(bh_gt1 / workday_count * 100, 2)  # 百分比，保留两位小数
        emp['下班一小时以上打卡率'] = f"{bh_rate}%" if bh_rate > 0 else ''
    else:
        emp['下班一小时以上打卡率'] = ''

    pr_ys = row['平日延时合计'] if pd.notna(row['平日延时合计']) else 0
    cc_dk_work = ew_orig[(ew_orig['出差合计'] > 0) &
                          (ew_orig['上班时间'].notna()) &
                          (ew_orig['下班时间'].notna())].shape[0]
    ps_gx = pr_ys + cc_days_work * 2 - cc_dk_work * 2
    emp['平时贡献'] = round(ps_gx, 2) if ps_gx != 0 else ''

    zl_ys = row['周六延时合计'] if pd.notna(row['周六延时合计']) else 0
    zr_ys = row['周日延时合计'] if pd.notna(row['周日延时合计']) else 0
    cc_dk_rest = er_orig[(er_orig['出差合计'] > 0) &
                          (er_orig['上班时间'].notna()) &
                          (er_orig['下班时间'].notna())].shape[0]
    zm_gx = zl_ys + zr_ys + cc_days_rest * 2 - cc_dk_rest * 2
    emp['周末贡献'] = round(zm_gx, 2) if zm_gx != 0 else ''

    jr_ys = row['假日延时合计'] if pd.notna(row['假日延时合计']) else 0
    emp['法定节假日带薪贡献'] = round(jr_ys, 2) if jr_ys != 0 else ''

    gx_sc = ps_gx + zm_gx + jr_ys
    emp['贡献时长'] = round(gx_sc, 2) if gx_sc != 0 else ''

    sj_days = row['产假'] if pd.notna(row['产假']) else 0
    cq_days = row['产前假'] if pd.notna(row['产前假']) else 0
    cj_days = row['产检假'] if pd.notna(row['产检假']) else 0
    mr_days = row['哺乳假'] if pd.notna(row['哺乳假']) else 0
    pc_days = row['陪产假'] if pd.notna(row['陪产假']) else 0
    bj_days = row['病假'] if pd.notna(row['病假']) else 0

    sq_days = sj_days + cq_days + cj_days + mr_days
    emp['三期天数'] = sq_days

    labels = []
    if sj_days > 0 or cq_days > 0 or cj_days > 0 or mr_days > 0:
        labels.append('三期')
    if pc_days > 0: labels.append('陪产')
    if bj_days > 0: labels.append('病假')
    emp['三期/陪产/病假'] = '/'.join(labels) if labels else ''

    if sj_days > 0 or cq_days > 0 or cj_days > 0 or mr_days > 0:
        emp['是否\n计入部门'] = '否'
    else:
        emp['是否\n计入部门'] = '是'

    is_problem = False
    bh_rate_val = bh_rate / 100 if workday_count > 0 and bh_rate > 0 else 0  # 转换为小数用于判断
    zm_rate_val = zm_rate / 100 if zm_rate > 0 else 0  # 转换为小数用于判断
    if gx_sc < 20 and zm_rate_val == 0 and bh_rate_val < 1/3 and dk_days >= 15 and cc_days_work == 0:
        is_problem = True
    emp['问题人员'] = '是' if is_problem else ''

    result_list.append(emp)

# ========== 生成结果 ==========
result_df = pd.DataFrame(result_list)
result_df['序号'] = range(1, len(result_df) + 1)

columns_order = ['序号', '浪潮工号', '职工姓名', '所属单位', '一级部门', '二级部门', '岗位', '考勤地',
                 '应出勤天数', '打卡天数', '出差天数（工作日）', '出差率（工作日）', '出差天数（休息日）',
                 '迟到次数', '请假天数\n（工作日）', '补签次数', '平均上班时间', '平均下班时间',
                 '周末出勤率', '下班一小时以上打卡率', '平时贡献', '周末贡献', '法定节假日带薪贡献',
                 '贡献时长', '三期天数', '三期/陪产/病假', '是否\n计入部门', '问题人员']

# 记录需要标黄的出差率位置（在删除临时列之前）
yellow_cells = set()
cc_rate_col_idx = columns_order.index('出差率（工作日）')
for row_idx in range(len(result_df)):
    if result_df.at[row_idx, '_出差率不确定']:
        yellow_cells.add((row_idx, cc_rate_col_idx))

result_df = result_df[columns_order]

# ========== 输出Excel并标黄 ==========
result_df.to_excel(OUTPUT_FILE, index=False)

wb = load_workbook(OUTPUT_FILE)
ws = wb.active
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 标黄（不确定的打卡地）
for (row_idx, col_idx) in yellow_cells:
    cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
    cell.fill = yellow_fill

wb.save(OUTPUT_FILE)

# ========== 统计输出 ==========
print(f"报表已保存到: {OUTPUT_FILE}（共{len(result_df)}人）")
if yellow_cells:
    print(f"出差率列有{len(yellow_cells)}个单元格标黄（打卡地识别不确定，需人工核实）")