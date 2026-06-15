from __future__ import annotations

import datetime
import json
import logging
import os
import re
from pathlib import Path
from typing import Any

import pandas as pd
import xlrd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def clean_text(s):
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    s = re.sub(r"[（(【\[]?单位[:：]\s*[^)】\]]+[)】\]]?", "", s)
    s = s.replace("→", "_").replace("￫", "_").replace("—", "_").replace("-", "_")
    s = s.replace("／", "_").replace("/", "_")
    s = re.sub(r"\s+", "_", s)
    return s.strip("_")


def looks_numeric_like(x):
    if x is None:
        return False
    if isinstance(x, (int, float)):
        return True
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return False
    s = s.replace(",", "").replace("%", "").replace("，", "").replace("％", "")
    try:
        float(s)
        return True
    except Exception:
        return False


def make_unique(names):
    seen = {}
    out = []
    for n in names:
        base = n or "col"
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
    return out


def coalesce_duplicate_columns(df):
    ordered_unique = []
    seen = set()
    for c in df.columns:
        if c not in seen:
            ordered_unique.append(c)
            seen.add(c)
    new_df = pd.DataFrame(index=df.index)
    for c in ordered_unique:
        g = df.loc[:, df.columns == c]
        coalesced = g.apply(lambda row: next((v for v in row.tolist() if pd.notna(v)), None), axis=1)
        new_df[c] = coalesced
    return new_df


def normalize_blank_strings(df):
    if df is None or not isinstance(df, pd.DataFrame):
        return df
    mask = df.map(lambda x: isinstance(x, str) and x.strip() == "")
    if mask.any(axis=None):
        df = df.where(~mask, None)
    return df


def parse_range_param(value):
    if value is None:
        return None
    if isinstance(value, int):
        if value <= 0:
            return None
        i = int(value) - 1
        return (i, i)
    if isinstance(value, list):
        if not value:
            return None
        value = value[0]
    s = str(value).strip()
    if s == "":
        return None
    if (s.startswith("'") and s.endswith("'")) or (s.startswith('"') and s.endswith('"')):
        s = s[1:-1].strip()
    if "-" in s:
        parts = s.split("-", 1)
        try:
            a = int(parts[0].strip())
            b = int(parts[1].strip())
            if a <= 0 or b <= 0:
                return None
            start = min(a, b) - 1
            end = max(a, b) - 1
            return (start, end)
        except Exception:
            return None
    try:
        a = int(s)
        if a <= 0:
            return None
        i = a - 1
        return (i, i)
    except Exception:
        return None


def get_unique_filename(directory, filename):
    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename
    while os.path.exists(os.path.join(directory, new_filename)):
        new_filename = f"{base}({counter}){ext}"
        counter += 1
    return new_filename


def process_dataframe(df, base_name, sheet_name, tmpdir, task_uuid, original_filename, sub_name_range=None, header_range=None):
    logger.info("[%s] 开始处理 DataFrame: %s, 形状: %s", task_uuid, sheet_name, df.shape)

    if df.isna().all().all():
        logger.warning("[%s] DataFrame 全为空，跳过处理", task_uuid)
        return []

    nonempty_col_mask = ~df.isna().all(axis=0)
    tables = []
    col = 0
    ncols = df.shape[1]
    empty_col_streak = 0

    while col < ncols:
        if not nonempty_col_mask.iloc[col]:
            empty_col_streak += 1
            if empty_col_streak >= 5:
                logger.info("[%s] 连续空列超过5个，停止分割", task_uuid)
                break
            col += 1
            continue
        empty_col_streak = 0
        start = col
        while col < ncols and nonempty_col_mask.iloc[col]:
            col += 1
        end = col
        t = df.iloc[:, start:end]
        if not t.isna().all(axis=None):
            tables.append(t.copy())
            logger.info("[%s] 找到子表 %s: 列范围 %s-%s, 形状: %s", task_uuid, len(tables), start, end - 1, t.shape)

    if not tables:
        logger.info("[%s] 未找到有效子表", task_uuid)
        return []

    logger.info("[%s] 共找到 %s 个子表", task_uuid, len(tables))
    local_paths = []

    for idx, t in enumerate(tables, 1):
        logger.info("[%s] 处理子表 %s/%s", task_uuid, idx, len(tables))
        t_raw = t.copy()
        t = t.ffill(axis=0)
        t = t.dropna(how="all").reset_index(drop=True)
        if t.empty:
            logger.info("[%s] 子表 %s 在移除空行后为空，跳过", task_uuid, idx)
            continue

        if sub_name_range is None:
            table_name_raw = str(sheet_name or "") or f"表{idx}"
        else:
            s_idx, e_idx = sub_name_range
            s_idx = max(0, s_idx)
            e_idx = max(0, e_idx)
            if s_idx >= len(t_raw):
                s_idx = len(t_raw) - 1
            if e_idx >= len(t_raw):
                e_idx = len(t_raw) - 1
            found = None
            for r in range(s_idx, e_idx + 1):
                row = t_raw.iloc[r]
                for cell in row:
                    if pd.notna(cell) and str(cell).strip() != "":
                        found = str(cell).strip()
                        break
                if found:
                    break
            table_name_raw = found or f"表{idx}"

        table_name = clean_text(table_name_raw) or f"表{idx}"
        header_rows_list = []
        data_start_row = None

        if header_range is not None:
            s_idx, e_idx = header_range
            s_idx = max(0, s_idx)
            e_idx = max(0, e_idx)
            if s_idx >= len(t_raw):
                s_idx = len(t_raw) - 1
            if e_idx >= len(t_raw):
                e_idx = len(t_raw) - 1

            for r in range(s_idx, e_idx + 1):
                header_rows_list.append(t_raw.iloc[r])

            last_header_row_in_raw = e_idx
            data_start_row_raw = last_header_row_in_raw + 1
            data_start_row = 0
            if data_start_row_raw < len(t_raw):
                non_empty_count = 0
                for r in range(len(t_raw)):
                    if not t_raw.iloc[r].isna().all():
                        if r >= data_start_row_raw:
                            break
                        non_empty_count += 1
                data_start_row = non_empty_count
            else:
                data_start_row = len(t)
        else:
            for r in range(1, len(t)):
                row = t.iloc[r]
                if row.isna().all():
                    continue
                nonnull_cnt = int(row.notna().sum())
                num_like_cnt = sum(looks_numeric_like(x) for x in row)
                if nonnull_cnt > 0 and (num_like_cnt >= 2 or (nonnull_cnt >= 3 and num_like_cnt / nonnull_cnt >= 0.5)):
                    data_start_row = r
                    break
                header_rows_list.append(row)
            if data_start_row is None:
                data_start_row = len(header_rows_list) + 1

        if header_rows_list:
            header_df = pd.DataFrame(header_rows_list).ffill(axis=0)
            header_df = header_df.map(lambda x: clean_text(x) if pd.notna(x) else None)
            names = []
            for c in range(header_df.shape[1]):
                parts = [x for x in header_df.iloc[:, c].tolist() if x]
                deduped = []
                for p in parts:
                    if not deduped or deduped[-1] != p:
                        deduped.append(p)
                name = "_".join(deduped).strip("_")
                names.append(name or f"col_{c + 1}")
        else:
            if len(t) > 0:
                first_row = t.iloc[0]
                names = [clean_text(str(x)) if pd.notna(x) else f"col_{i + 1}" for i, x in enumerate(first_row)]
            else:
                names = [f"col_{c + 1}" for c in range(t.shape[1])]
            base_original_name = os.path.splitext(str(original_filename))[0]
            safe_table_name = re.sub(r"[\\\/\?\*\[\]\:]", "_", table_name) or f"table_{idx}"
            names = make_unique([clean_text(n) for n in names])
            names = [f"{base_original_name}__{safe_table_name}_{col}" for col in names]

        df_data = t.iloc[data_start_row:].reset_index(drop=True)
        non_all_nan_cols = ~df_data.isna().all(axis=0)

        if df_data.shape[1] != len(names):
            names = names[:df_data.shape[1]] + [f"col_{i + 1}" for i in range(len(names), df_data.shape[1])]

        df_data = df_data.loc[:, non_all_nan_cols]
        names = [n for n, keep in zip(names, non_all_nan_cols) if keep]

        if df_data.shape[1] == 0 or df_data.dropna(how="all").empty:
            continue

        names = make_unique([clean_text(n) for n in names])
        df_data.columns = names
        df_data = coalesce_duplicate_columns(df_data)
        df_data = df_data.ffill(axis=0)
        df_data = df_data.dropna(how="all")
        df_data = df_data.loc[:, ~df_data.isna().all(axis=0)]
        df_data = df_data.drop_duplicates()

        if df_data.empty or df_data.shape[1] == 0:
            continue

        base_original_name = os.path.splitext(str(original_filename))[0]
        safe_table_name = re.sub(r"[\\\/\?\[\]\:]", "_", table_name) if table_name else f"table_{idx}"
        safe_table_name = safe_table_name or f"table_{idx}"
        flattened_headers = [f"{base_original_name}__{safe_table_name}_{col}" for col in df_data.columns]
        df_data.columns = make_unique(flattened_headers)

        candidate_name = f"{base_name}__{safe_table_name}"[:120]
        final_filename = get_unique_filename(tmpdir, f"{candidate_name}.xlsx")
        temp_output_file = os.path.join(tmpdir, final_filename)
        df_data.to_excel(temp_output_file, index=False)
        local_paths.append(temp_output_file)

    if len(local_paths) == 0 and len(tables) > 0:
        t = tables[0].dropna(how="all").reset_index(drop=True)
        if t.empty:
            return local_paths
        if len(t) > 1:
            raw_headers = t.iloc[0]
            cleaned_headers = [clean_text(x) for x in raw_headers]
            unique_headers = make_unique(cleaned_headers)
            base_original_name = os.path.splitext(str(original_filename))[0]
            safe_table_name = re.sub(r"[\\\/\?\[\]\:]", "_", table_name) or f"table_{idx}"
            flattened_headers = [f"{base_original_name}__{safe_table_name}_{col}" for col in unique_headers]
            t.columns = make_unique(flattened_headers)
            t = t.iloc[1:].reset_index(drop=True)
        else:
            base_original_name = os.path.splitext(str(original_filename))[0]
            safe_table_name = re.sub(r"[\\\/\?\[\]\:]", "_", table_name) or f"table_{idx}"
            t.columns = [f"{base_original_name}__{safe_table_name}_col_{i + 1}" for i in range(t.shape[1])]

        t = coalesce_duplicate_columns(t)
        t = t.dropna(how="all")
        t = t.loc[:, ~t.isna().all(axis=0)]
        t = t.drop_duplicates()

        if t.empty or t.shape[1] == 0:
            return local_paths

        candidate_name = f"{os.path.splitext(original_filename)[0]}__{safe_table_name}"[:120]
        final_filename = get_unique_filename(tmpdir, f"{candidate_name}.xlsx")
        temp_output_file = os.path.join(tmpdir, final_filename)
        t.to_excel(temp_output_file, index=False)
        local_paths.append(temp_output_file)

    return local_paths


def process_file(input_path, tmpdir, task_uuid, original_filename, sub_name_range=None, header_range=None):
    logger.info("[%s] 开始处理文件: %s -> %s", task_uuid, original_filename, input_path)
    ext = str(input_path).split(".")[-1].lower()
    base_name = clean_text(os.path.splitext(os.path.basename(str(input_path)))[0])
    all_local_paths = []

    def first_nonempty_in_grid(grid, r1, r2, c1, c2):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                try:
                    val = grid[rr][cc]
                except Exception:
                    val = None
                if val is None:
                    continue
                if isinstance(val, str) and val.strip() == "":
                    continue
                return val
        return None

    def apply_row_left_to_right_fill(grid, merged_mask=None):
        if merged_mask is None:
            return
        for r in range(len(grid)):
            prev = None
            for c in range(len(grid[r])):
                val = grid[r][c]
                is_empty = val is None or (isinstance(val, str) and str(val).strip() == "")
                if is_empty:
                    if prev is not None and merged_mask[r][c]:
                        grid[r][c] = prev
                    else:
                        grid[r][c] = None
                else:
                    prev = grid[r][c]

    if ext == "xlsx":
        wb = load_workbook(str(input_path), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_r, max_c = ws.max_row, ws.max_column
            grid = [[None for _ in range(max_c)] for _ in range(max_r)]
            merged_mask = [[False for _ in range(max_c)] for _ in range(max_r)]
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(r, c)
                    val = cell.value
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        try:
                            grid[r - 1][c - 1] = f"{val.year}/{val.month}/{val.day}"
                        except Exception:
                            grid[r - 1][c - 1] = str(val)
                    else:
                        grid[r - 1][c - 1] = val
            if hasattr(ws, "merged_cells") and getattr(ws, "merged_cells") is not None:
                for merged in getattr(ws, "merged_cells").ranges:
                    try:
                        min_row = merged.min_row
                        max_row = merged.max_row
                        min_col = merged.min_col
                        max_col = merged.max_col
                        first_val = first_nonempty_in_grid(grid, min_row - 1, max_row - 1, min_col - 1, max_col - 1)
                        for rr in range(min_row - 1, max_row):
                            for cc in range(min_col - 1, max_col):
                                if 0 <= rr < max_r and 0 <= cc < max_c:
                                    merged_mask[rr][cc] = True
                        if first_val is not None:
                            for rr in range(min_row - 1, max_row):
                                for cc in range(min_col - 1, max_col):
                                    if 0 <= rr < max_r and 0 <= cc < max_c:
                                        grid[rr][cc] = first_val
                        if max_row == min_row and max_col > min_col and min_row < max_r:
                            for cc in range(min_col, max_col + 1):
                                r0 = min_row - 1
                                c0 = cc - 1
                                if 0 <= r0 < max_r and 0 <= c0 < max_c and r0 + 1 < max_r:
                                    current_val = grid[r0][c0]
                                    next_row_val = grid[r0 + 1][c0]
                                    if current_val is not None and next_row_val is not None:
                                        grid[r0][c0] = f"{current_val}_{next_row_val}"
                                    elif current_val is not None:
                                        grid[r0][c0] = current_val
                                    else:
                                        grid[r0][c0] = next_row_val
                    except Exception:
                        continue
            apply_row_left_to_right_fill(grid, merged_mask=merged_mask)
            df = normalize_blank_strings(pd.DataFrame(grid))
            result = process_dataframe(df, base_name, sheet_name, tmpdir, task_uuid, original_filename, sub_name_range, header_range)
            if isinstance(result, list):
                all_local_paths.extend(result)
            elif result:
                all_local_paths.append(result)
    elif ext == "xls":
        try:
            book = xlrd.open_workbook(str(input_path), formatting_info=True)
        except Exception:
            book = xlrd.open_workbook(str(input_path))
        for sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            rows = sheet.nrows
            cols = sheet.ncols
            grid = [[None for _ in range(cols)] for _ in range(rows)]
            merged_mask = [[False for _ in range(cols)] for _ in range(rows)]
            for r in range(rows):
                for c in range(cols):
                    ctype = sheet.cell_type(r, c)
                    raw_val = sheet.cell_value(r, c)
                    if ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(raw_val, book.datemode)
                            grid[r][c] = f"{dt.year}/{dt.month}/{dt.day}"
                        except Exception:
                            try:
                                grid[r][c] = str(raw_val)
                            except Exception:
                                grid[r][c] = None
                    else:
                        grid[r][c] = raw_val
            if hasattr(sheet, "merged_cells"):
                for (rlo, rhi, clo, chi) in sheet.merged_cells:
                    try:
                        first_val = first_nonempty_in_grid(grid, rlo, rhi - 1, clo, chi - 1)
                        for rr in range(rlo, rhi):
                            for cc in range(clo, chi):
                                if 0 <= rr < rows and 0 <= cc < cols:
                                    merged_mask[rr][cc] = True
                        if first_val is not None:
                            for rr in range(rlo, rhi):
                                for cc in range(clo, chi):
                                    if 0 <= rr < rows and 0 <= cc < cols:
                                        grid[rr][cc] = first_val
                        if rhi == rlo + 1 and chi > clo + 1 and rlo < rows - 1:
                            for cc in range(clo, chi):
                                if 0 <= rlo < rows and 0 <= cc < cols:
                                    current_val = grid[rlo][cc]
                                    next_row_val = grid[rlo + 1][cc] if (rlo + 1) < rows else None
                                    if current_val is not None and next_row_val is not None:
                                        grid[rlo][cc] = f"{current_val}_{next_row_val}"
                                    elif current_val is not None:
                                        grid[rlo][cc] = current_val
                                    else:
                                        grid[rlo][cc] = next_row_val
                    except Exception:
                        continue
            apply_row_left_to_right_fill(grid, merged_mask=merged_mask)
            df = normalize_blank_strings(pd.DataFrame(grid))
            result = process_dataframe(df, base_name, sheet_name, tmpdir, task_uuid, original_filename, sub_name_range, header_range)
            if isinstance(result, list):
                all_local_paths.extend(result)
            elif result:
                all_local_paths.append(result)
    elif ext == "csv":
        try:
            df = pd.read_csv(str(input_path), header=None, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(str(input_path), header=None, dtype=str, encoding="gbk", errors="ignore")
        df2 = normalize_blank_strings(df.fillna("").replace("", None))
        result = process_dataframe(df2, base_name, "csv_sheet", tmpdir, task_uuid, original_filename, sub_name_range, header_range)
        if isinstance(result, list):
            all_local_paths.extend(result)
        elif result:
            all_local_paths.append(result)
    elif ext == "json":
        df = pd.read_json(str(input_path))
        df2 = normalize_blank_strings(df.fillna("").replace("", None))
        result = process_dataframe(df2, base_name, "json_sheet", tmpdir, task_uuid, original_filename, sub_name_range, header_range)
        if isinstance(result, list):
            all_local_paths.extend(result)
        elif result:
            all_local_paths.append(result)
    else:
        raise Exception(f"不支持的文件格式: {ext}")

    return all_local_paths


def get_file_metadata(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".csv":
            df = pd.read_csv(file_path, nrows=3)
        elif ext in [".xls", ".xlsx"]:
            df = pd.read_excel(file_path, nrows=3)
        elif ext == ".json":
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                df = pd.DataFrame(data).head(3)
            else:
                df = pd.DataFrame([data]).head(3)
        else:
            return None

        def clean_json_data(obj):
            if isinstance(obj, dict):
                return {k: clean_json_data(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [clean_json_data(item) for item in obj]
            if isinstance(obj, float):
                if pd.isna(obj):
                    return None
                if obj == float("inf"):
                    return "Infinity"
                if obj == float("-inf"):
                    return "-Infinity"
                return obj
            return obj

        sample_data = df.to_dict(orient="records")
        cleaned_sample_data = clean_json_data(sample_data)
        return {
            "columns": list(df.columns),
            "sample_data": cleaned_sample_data,
            "file_name": os.path.basename(file_path),
        }
    except Exception:
        return None
